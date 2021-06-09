#reading values
import openpyxl

book = openpyxl.load_workbook('test.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['B1']
a3 = sheet.cell(row=3, column=1)

print(a1.value)
print(a2.value) 
print(a3.value)

#writing values 1
from openpyxl import Workbook
import time

book = Workbook()
sheet = book.active

sheet['A1'] = 56
sheet['A2'] = 43

now = time.strftime("%x")
sheet['A3'] = now

book.save("sample.xlsx")

#writing values 2
book = Workbook()
sheet = book.active

sheet['A1'] = 1
sheet.cell(row=2, column=2).value = 2

book.save('./excel_output/write2cell.xlsx')

#appending
#from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

book.save('./excel_output/appending.xlsx')

#iterate by rows
#from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)
    
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()    

book.save('./excel_output/iterbyrows.xlsx')

# iterate by columns
# from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)
    
for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()    

book.save('./excel_output/iterbycols.xlsx')

#filter/sort
#from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

data = [
    ['Item', 'Colour'],
    ['pen', 'brown'],
    ['book', 'black'],
    ['plate', 'white'],
    ['chair', 'brown'],
    ['coin', 'gold'],
    ['bed', 'brown'],
    ['notebook', 'white'],
]

for r in data:
    sheet.append(r)

sheet.auto_filter.ref = 'A1:B8'
sheet.auto_filter.add_filter_column(1, ['brown', 'white'])
sheet.auto_filter.add_sort_condition('B2:B8')

wb.save('filtered.xlsx')

#dimensions
#from openpyxl import Workbook

book = Workbook()
sheet = book.active

sheet['A3'] = 39
sheet['B3'] = 19

rows = [
    (88, 46),
    (89, 38),
    (23, 59),
    (56, 21),
    (24, 18),
    (34, 15)
]

for row in rows:
    sheet.append(row)

print(sheet.dimensions)
print("Minimum row: {0}".format(sheet.min_row))
print("Maximum row: {0}".format(sheet.max_row))
print("Minimum column: {0}".format(sheet.min_column))
print("Maximum column: {0}".format(sheet.max_column))

for c1, c2 in sheet[sheet.dimensions]:
    print(c1.value, c2.value)

book.save('./excel_output/dimensions.xlsx')

#merging
#from openpyxl import Workbook
from openpyxl.styles import Alignment

book = Workbook()
sheet = book.active

sheet.merge_cells('A1:B2')

cell = sheet.cell(row=1, column=1)
cell.value = 'Sunny day'
cell.alignment = Alignment(horizontal='center', vertical='center')

book.save('./excel_output/merging.xlsx')

#formulas
#from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (34, 26),
    (88, 36),
    (24, 29),
    (15, 22),
    (56, 13),
    (76, 18)
)

for row in rows:
    sheet.append(row)

cell = sheet.cell(row=7, column=2)
cell.value = "=SUM(A1:B6)"
cell.font = cell.font.copy(bold=True)

book.save('./excel_output/formulas.xlsx')

#charts
#from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)

book = Workbook()
sheet = book.active

rows = [
    ("USA", 46),
    ("China", 38),
    ("UK", 29),
    ("Russia", 22),
    ("South Korea", 13),
    ("Germany", 11)
]

for row in rows:
    sheet.append(row)
    
data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=6)
categs = Reference(sheet, min_col=1, min_row=1, max_row=6)

chart = BarChart()
chart.add_data(data=data)
chart.set_categories(categs)

chart.legend = None
chart.y_axis.majorGridlines = None
chart.varyColors = True
chart.title = "Olympic Gold medals in London"

sheet.add_chart(chart, "A8")    

book.save("bar_chart.xlsx")